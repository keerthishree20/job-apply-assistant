from datetime import date
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

EXCEL_PATH = Path("jobs_applied.xlsx")
HEADERS = ["S.No", "Company", "Role", "Job URL", "Date Applied",
           "Status", "Resume Version", "Cover Letter", "Notes"]


def _create_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["B"].width = 25
    return wb


def append_application(
    company: str,
    role: str,
    url: str,
    cover_letter: bool = True,
    resume_snippet: str = "",
    notes: str = "",
) -> None:
    wb = load_workbook(EXCEL_PATH) if EXCEL_PATH.exists() else _create_workbook()
    ws = wb.active
    s_no = ws.max_row  # header is row 1, so max_row gives next S.No
    ws.append([
        s_no,
        company,
        role,
        url,
        date.today().strftime("%d-%m-%Y"),
        "Applied",
        resume_snippet[:120] if resume_snippet else "",
        "Yes" if cover_letter else "No",
        notes,
    ])
    wb.save(EXCEL_PATH)


def get_excel_path() -> Path:
    if not EXCEL_PATH.exists():
        wb = _create_workbook()
        wb.save(EXCEL_PATH)
    return EXCEL_PATH
